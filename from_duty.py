import openpyxl
import datetime


def get_cell_x_num_by_letter(letter):
    if ord(letter) >= ord('A') and ord(letter) <= ord('Z'):
        return ord(letter) - ord('A') + 1
    elif ord(letter) >= ord('a') and ord(letter) <= ord('z'):
        return ord(letter) - ord('a') + 1
    else:
        raise Exception("<Error> Wrong letter {}".format(letter))


def get_time_of_day_str(tod):
    if tod == 0:
        return "Завтрак"
    elif tod == 1:
        return "Обед"
    elif tod == 2:
        return "Ужин"
    else:
        raise Exception("<Error> wrong time of date {}".format(tod))


class FoodTime(object):
    def __init__(self, person, date, time_of_day, general_food_type):
        self.person = person
        self.date = date
        self.time_of_day = time_of_day
        self.general_food_type = general_food_type

    def __repr__(self):
        return "{} {} {} {}".format(self.date.strftime('%Y-%m-%d'), self.person, self.time_of_day, self.general_food_type)


# Словарь с расположением типов еды на листе
MIEL_PLACE = {
    "Геркулес": "G8", "Пшенка": "I8",
    "Харчо": "G14", "Борщ": "I14", "Гороховый": "K14",
    "Гречка": "G20", "Рис": "I20", "Макароны": "K20"
}


file_name = "Раскладка. Питание _24.05.04.xlsx"
wb = openpyxl.load_workbook(file_name, read_only=False, data_only=True)  # Открыть таблицу с данными о НС
sheet_names = wb.sheetnames
print("Найденные листы в книге: {}".format(sheet_names))
sheet_name = "График"
ws = wb[sheet_name]

#some = ws.cell(row=10, column=6).value
#print(some)
#quit()

start_x = get_cell_x_num_by_letter('F')
end_x = get_cell_x_num_by_letter('N')
print("start_x = {}, end_x = {}".format(start_x, end_x))
meal_x = get_cell_x_num_by_letter('E')

start_y = 8
end_y = 34

all_count = 0

data_base = []
start_date = datetime.datetime(year=2024, month=5, day=4)
for x in range(start_x, end_x + 1):
    date = start_date + datetime.timedelta(days=x - 1)
    for y in range(start_y, end_y + 1):
        person = ws.cell(row=y, column=x).value
        if person is None:
            continue
        print("row = {} col = {} val = {}".format(y, x, person))
        time_of_day = (y - start_y) % 3
        meal = ws.cell(row=y, column=meal_x).value
        data_base.append(FoodTime(
            person=person,
            date=date,
            time_of_day=get_time_of_day_str(time_of_day),
            general_food_type=meal
        ))
        all_count += 1

test_count = int(ws['P6'].value)  # Проверяем данные

print("all_count = {}, test_count = {}".format(all_count, test_count))
if all_count != test_count:
    raise Exception("all_count != test_count {} != {}".format(all_count, test_count))

i = 0
for d in data_base:
    print("{}. {}".format(i, d))
    i += 1

persons = set([ft.person for ft in data_base])
target = wb.copy_worksheet(wb["Шаблон"])
print(persons)