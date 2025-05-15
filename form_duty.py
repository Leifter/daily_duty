import openpyxl
import datetime

"""
Скрипт для формирования раскладок в соответствии с графиком дежурств. Автор Панин С.А.
Как пользоваться
!!! Внимание !!! Не допускаются изменения структуры шаблона, без согласования его с данным скриптом. Так, нельзя менять
 формулы и добавлять, удалять строки или столбцы, так как это приведет к изменению координат, заданных в скрипте
1. Заполнить шаблон в файле "Исходный_Шаблон.xlsx"
1.1 На листе "График"
    - Заполнить версию на графике
    - Заполнить количество участников на каждый день
    - Заполнить кто, когда дежурит
    - Заполнить базовые блюда в колонке блюд
    - Заполнить количество персон на каждом приеме пищи
    - Провести пересчет ячейки "Общее кол-во приемов пищи". Убедиться, что рассчет захватывает все требуемые ячейки. 
    Это число требуется только для рассчета количества дежурств на одного человека и процедуры проверки в скрипте
    - Если требуется, добавить по аналогии новые блюда
1.2 На листе "Шаблон"
    - Заполнить типы супов и согласовать их с общей суммой (~E14), из которой высчитывается количество обедов
    - Заполнить типы каш с молоком и согласовать их с общей суммой (~E8), из которой высчитывается количество завтраков или ужинов  
    - Заполнить типы каш с мясом и согласовать их с общей суммой (~E20), из которой высчитывается количество завтраков или ужинов
    - Если требуется поменять каши с мясом и каши с молоком, то заменить 1. Завтрак на 5 Ужин, при этом остальные ячейки не перемещать
    - Добавить или удалить рассчет перловки, как каши, добавляемые к какой-то базовой
    - Творчески подойти к редактированию данной страницы 
2. Согласовать скрипт с шаблоном
    - Заполнить словарь MEAL_PLACE в соответствии с координатами количества блюд
    - Заполнить переменные START_X, START_Y, END_X, END_Y в соответствии размерами таблицы дежурств
    - Заполнить переменную out_file_name в соответствии с именем итогового файла дежурств
3. Выполнить скрипт python form_duty.py. Естественно для этого потребуется установленный python и пакет openpyxl
4a. Если скрипт выполнился без ошибок вручную проверить итоговую раскладку. Если обнаружены ошибки обратиться к 
    разработчику за разъяснениями
4b. Если скрипт выполнился с ошибками обратиться к разработчику за разъяснениями
"""

# ToDo
#  - 1. Добавить проверку наличия данного типа еды в графике в колонке с едой
#  2. Выделить в отдельные столбцы штуки и граммы, с тем чтобы можно было считать суммарную массу

def get_cell_letter_num_by_x(n):
    """Преобразует номер столбца в буквенное обозначение Excel"""
    result = ""
    if n <= 0:
        raise Exception("n <= 0 n = {}".format(n))
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

def get_cell_x_num_by_letter(column):
    """
    Преобразует буквенное обозначение Excel-столбца (например, 'A', 'Z', 'AA', 'AZ') в номер столбца (1, 26, 27, 52).
    """
    column = column.upper()
    num = 0
    for c in column:
        if ord(c) < ord('A') or ord(c) > ord('Z'):
            raise Exception("<Error> Wrong letter {}".format(c))

        num = num * 26 + (ord(c) - ord('A') + 1)
    return num

def get_cell_x_num_by_letter_old(letter):
    if ord(letter) >= ord('A') and ord(letter) <= ord('Z'):
        return ord(letter) - ord('A') + 1
    elif ord(letter) >= ord('a') and ord(letter) <= ord('z'):
        return ord(letter) - ord('a') + 1
    else:
        raise Exception("<Error> Wrong letter {}".format(letter))


def get_time_of_day_str(tod):
    if tod == 0:
        return "Завтрак"
    elif tod == 1:
        return "Обед"
    elif tod == 2:
        return "Ужин"
    else:
        raise Exception("<Error> wrong time of date {}".format(tod))


class FoodTime(object):
    def __init__(self, person, date, time_of_day, general_food_type, meal_portions):
        self.person = person
        self.date = date
        self.time_of_day = time_of_day
        self.general_food_type = general_food_type
        self.meal_portions = meal_portions

    def __repr__(self):
        return "{} {} {} {} {} порций".format(
            self.date.strftime('%Y-%m-%d'),
            self.person,
            self.time_of_day,
            self.general_food_type,
            self.meal_portions
        )


# Словарь с расположением типов еды на листе
MEAL_PLACE = {
    "Геркулес": "G8", "Пшенка": "I8",
    "Харчо": "G14", "Борщ": "I14", "Гороховый": "K14",
    "Гречка": "G20", "Рис": "I20", "Макароны": "K20", "Без супа": "M14"
}

def get_meal_count(patten_file):
    """
    Отдельно открываем график, чтобы считать общее количество питаний, так как оно рассчитывается
    , а openpyxl сам по себе вычислять формулы не умеет
    :param patten_file:
    :return:
    """
    wb = openpyxl.load_workbook(patten_file, data_only=True)  # Открыть таблицу с данными о НС
    sheet_names = wb.sheetnames
    print("Найденные листы в книге: {}".format(sheet_names))
    sheet_name = "График"
    ws = wb[sheet_name]
    all_meal_count = int(ws['B6'].value)  # Проверяем данные
    wb.close()
    return all_meal_count

def form_duty(patten_file, out_file_name):

    test_count = get_meal_count(patten_file)

    wb = openpyxl.load_workbook(patten_file, read_only=False)  # Открыть таблицу с данными о НС
    sheet_names = wb.sheetnames
    print("Найденные листы в книге: {}".format(sheet_names))
    sheet_name = "График"
    ws = wb[sheet_name]

    #some = ws.cell(row=10, column=6).value
    #print(some)
    #quit()

    START_X = get_cell_x_num_by_letter('G')   # X-координата левого верхнего угла расписания, с которого начинаются фамилии дежурных
    START_Y = 11                              # Y-координата левого верхнего угла расписания

    END_X = get_cell_x_num_by_letter('AC')    # X-координата правого нижнего угла расписания, на котором заканчиваются фамилии дежурных
    END_Y = 77                                # Y-координата правого нижнего угла расписания

    print("START_X = {}, end_x = {}".format(START_X, END_X))
    meal_x = get_cell_x_num_by_letter('E')        # Колонка с блюдом
    meal_count_x = get_cell_x_num_by_letter('D')  # Колонка с количеством персон

    all_count = 0

    data_base = []
    start_date = datetime.datetime(year=2024, month=5, day=4)
    for x in range(START_X, END_X + 1):
        date = start_date + datetime.timedelta(days=x - 1)
        for y in range(START_Y, END_Y + 1):
            person = ws.cell(row=y, column=x).value
            if person is None:
                continue
            print("row = {} col = {} val = {}".format(y, x, person))
            time_of_day = (y - START_Y) % 3
            meal = ws.cell(row=y, column=meal_x).value
            # Проверка наличия данного типа еды в списке MEAL_PLACE
            if meal not in MEAL_PLACE.keys():
                raise Exception("meal {} at cell {}{} not in MEAL_PLACE = {}".format(
                    meal,
                    get_cell_letter_num_by_x(meal_x),
                    y,
                    MEAL_PLACE.keys()))

            meal_portions = ws.cell(row=y, column=meal_count_x).value
            data_base.append(FoodTime(
                person=person,
                date=date,
                time_of_day=get_time_of_day_str(time_of_day),
                general_food_type=meal,
                meal_portions=meal_portions
            ))
            all_count += 1

    print("all_count = {}, test_count = {}".format(all_count, test_count))
    if all_count != test_count:
        raise Exception("all_count != test_count {} != {}".format(all_count, test_count))

    i = 0
    for d in data_base:
        print("{}. {}".format(i, d))
        i += 1

    persons = set([ft.person for ft in data_base])   # Получаем имена всех участников
    persons = sorted(persons)
    persons_meals = {}
    for p in persons:
        target = wb.copy_worksheet(wb["Шаблон"])
        target.title = p
        persons_meals[p] = {}
        for ft in data_base:
            if ft.person == p:
                try:
                    persons_meals[p][ft.general_food_type] += ft.meal_portions
                except KeyError:
                    persons_meals[p][ft.general_food_type] = ft.meal_portions

        for meal in persons_meals[p]:
            try:
                target[MEAL_PLACE[meal]] = persons_meals[p][meal]
            except KeyError:
                raise Exception("В списке отсутствует еда {}".format(meal))


    print(persons_meals)

    wb.save(out_file_name)


if __name__ == "__main__":
    if get_cell_x_num_by_letter('D') != get_cell_x_num_by_letter_old('D'):
        raise Exception("Refactoring error")

    form_duty(patten_file="Исходный_Шаблон.xlsx", out_file_name="Раскладка_Питание_2025.05.23.xlsx")


